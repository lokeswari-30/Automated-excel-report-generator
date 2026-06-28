import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

df=pd.read_csv("sample_data.csv")
summary={"Average":df["Marks"].mean(),"Highest":df["Marks"].max(),"Lowest":df["Marks"].min()}
out="output/generated_report.xlsx"
with pd.ExcelWriter(out,engine="openpyxl") as writer:
    df.to_excel(writer,index=False,sheet_name="Report")
    pd.DataFrame(summary,index=[0]).to_excel(writer,index=False,sheet_name="Summary")
wb=load_workbook(out)
for ws in wb.worksheets:
    for c in ws[1]:
        c.font=Font(bold=True)
wb.save(out)
print("Report generated:",out)
